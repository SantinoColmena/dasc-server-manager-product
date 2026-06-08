# -*- coding: utf-8 -*-
"""Genera el Excel derivado del roadmap de Vigex.

Uso (desde la raíz del repo o desde cualquier sitio):

    python tools/roadmap/generar_roadmap_xlsx.py

Lee los datos de ``roadmap_data.py`` y escribe
``docs/roadmap/Vigex_Roadmap.xlsx``. El Excel es una VISTA DERIVADA: no se edita
a mano, se regenera. La fuente de verdad narrativa es ``docs/ROADMAP.md``.
"""

from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Permite importar roadmap_data tanto si se ejecuta desde la raíz como desde aquí.
sys.path.insert(0, str(Path(__file__).resolve().parent))
import roadmap_data as data  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parents[2]
SALIDA = REPO_ROOT / "docs" / "roadmap" / "Vigex_Roadmap.xlsx"

# ----------------------------------------------------------------------- estilo
AZUL = "1F3864"
AZUL_CLARO = "2E5496"
GRIS_CAB = "D9D9D9"
BLANCO = "FFFFFF"

COLOR_ESTADO = {
    "Cerrada": "C6EFCE",
    "En curso": "BDD7EE",
    "Siguiente": "9FE2E2",
    "Planificada": "FFF2CC",
    "Backlog": "E7E6E6",
    "Aplazada": "FCE4D6",
    "Diferida": "D9D9D9",
}

FUENTE_CAB = Font(bold=True, color=BLANCO, size=11)
RELLENO_CAB = PatternFill("solid", fgColor=AZUL)
RELLENO_TITULO = PatternFill("solid", fgColor=AZUL)
BORDE_FINO = Border(*[Side(style="thin", color="BFBFBF")] * 4)
CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
IZQ = Alignment(horizontal="left", vertical="center", wrap_text=True)
IZQ_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _cab(ws, fila, valores):
    """Escribe una fila de cabecera con estilo."""
    for c, val in enumerate(valores, start=1):
        cel = ws.cell(row=fila, column=c, value=val)
        cel.font = FUENTE_CAB
        cel.fill = RELLENO_CAB
        cel.alignment = CENTRO
        cel.border = BORDE_FINO


def _anchos(ws, anchos):
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------------- hoja: Léeme
def hoja_leeme(wb):
    ws = wb.active
    ws.title = "Léeme"
    _anchos(ws, [22, 95])
    ws["A1"] = f"{data.META['producto']} — Hoja de ruta"
    ws["A1"].font = Font(bold=True, size=16, color=AZUL)
    ws.merge_cells("A1:B1")

    filas = [
        ("Versión", data.META["version"]),
        ("Fase actual", data.META["fase_actual"]),
        ("Actualizado", data.META["actualizado"]),
        ("Fuente de verdad", data.META["fuente"]),
        ("", ""),
        ("IMPORTANTE", "Este Excel se GENERA. No editar a mano: cambiar docs/ROADMAP.md "
                       "y tools/roadmap/roadmap_data.py, y regenerar."),
        ("Regenerar", "python tools/roadmap/generar_roadmap_xlsx.py"),
        ("", ""),
        ("Hojas", "Roadmap = tabla maestra · Gates = criterios de salida · "
                  "Perfiles = arquitecturas · Decisiones = registro · Resumen = KPIs"),
        ("", ""),
        ("Jerarquía", "Fase -> Ruta -> Tarea (R-xxx) -> Gate. Subrutas con sufijo "
                      "de letra (R-049A). Los R-xxx no se reutilizan ni se renumeran."),
    ]
    r = 3
    for k, v in filas:
        ca = ws.cell(row=r, column=1, value=k)
        ca.font = Font(bold=True, color=AZUL_CLARO)
        ca.alignment = IZQ_TOP
        cb = ws.cell(row=r, column=2, value=v)
        cb.alignment = IZQ_TOP
        r += 1

    # Leyenda de estados.
    r += 1
    ws.cell(row=r, column=1, value="Leyenda de estados").font = Font(bold=True, color=AZUL)
    r += 1
    for estado, color in COLOR_ESTADO.items():
        cel = ws.cell(row=r, column=1, value=estado)
        cel.fill = PatternFill("solid", fgColor=color)
        cel.alignment = CENTRO
        cel.border = BORDE_FINO
        r += 1


# ------------------------------------------------------------------- hoja: Roadmap
def hoja_roadmap(wb):
    ws = wb.create_sheet("Roadmap")
    cols = data.COLUMNAS
    _cab(ws, 1, cols)
    _anchos(ws, [9, 34, 18, 46, 11, 12, 11, 12, 60, 42])

    for i, fila in enumerate(data.ROADMAP, start=2):
        es_fase = fila["Tipo"] == "Fase"
        es_sub = fila["Tipo"] == "Subtarea"
        for c, nombre in enumerate(cols, start=1):
            val = fila.get(nombre, "")
            if nombre == "Título" and es_sub:
                val = "    " + val  # sangría visual de subtareas
            cel = ws.cell(row=i, column=c, value=val)
            cel.border = BORDE_FINO
            cel.alignment = CENTRO if nombre in ("ID", "Tipo", "Estado", "Horizonte", "Gate") else IZQ_TOP

            if nombre == "Estado":
                color = COLOR_ESTADO.get(val)
                if color:
                    cel.fill = PatternFill("solid", fgColor=color)

        # Resalta filas de Fase.
        if es_fase:
            for c in range(1, len(cols) + 1):
                cel = ws.cell(row=i, column=c)
                cel.fill = RELLENO_TITULO
                cel.font = Font(bold=True, color=BLANCO)
                if cols[c - 1] == "Estado":  # mantiene color de estado legible
                    cel.font = Font(bold=True, color="000000")
                    color = COLOR_ESTADO.get(fila["Estado"])
                    if color:
                        cel.fill = PatternFill("solid", fgColor=color)
        elif es_sub:
            for c in range(1, len(cols) + 1):
                ws.cell(row=i, column=c).font = Font(color="404040", italic=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data.ROADMAP) + 1}"


# --------------------------------------------------------------------- hoja: Gates
def hoja_gates(wb):
    ws = wb.create_sheet("Gates")
    cab = ["Gate", "Nombre", "Estado", "Criterio de salida (real)", "Evidencia"]
    _cab(ws, 1, cab)
    _anchos(ws, [14, 32, 12, 80, 26])
    for i, (gate, nombre, estado, criterio, ev) in enumerate(data.GATES, start=2):
        valores = [gate, nombre, estado, criterio, ev]
        for c, val in enumerate(valores, start=1):
            cel = ws.cell(row=i, column=c, value=val)
            cel.border = BORDE_FINO
            cel.alignment = CENTRO if c in (1, 3) else IZQ_TOP
            if c == 3:
                color = COLOR_ESTADO.get("Cerrada" if estado == "Cerrada" else "Planificada")
                cel.fill = PatternFill("solid", fgColor=color)
    ws.freeze_panes = "A2"


# ------------------------------------------------------------------ hoja: Perfiles
def hoja_perfiles(wb):
    ws = wb.create_sheet("Perfiles")
    cab = ["Perfil", "Arquitectura", "Uso recomendado", "Decisión"]
    _cab(ws, 1, cab)
    _anchos(ws, [16, 40, 34, 42])
    for i, fila in enumerate(data.PERFILES, start=2):
        for c, val in enumerate(fila, start=1):
            cel = ws.cell(row=i, column=c, value=val)
            cel.border = BORDE_FINO
            cel.alignment = IZQ_TOP
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------- hoja: Decisiones
def hoja_decisiones(wb):
    ws = wb.create_sheet("Decisiones")
    cab = ["Fecha", "Decisión", "Motivo"]
    _cab(ws, 1, cab)
    _anchos(ws, [14, 70, 55])
    for i, fila in enumerate(data.DECISIONES, start=2):
        for c, val in enumerate(fila, start=1):
            cel = ws.cell(row=i, column=c, value=val)
            cel.border = BORDE_FINO
            cel.alignment = CENTRO if c == 1 else IZQ_TOP
    ws.freeze_panes = "A2"


# ------------------------------------------------------------------- hoja: Resumen
def hoja_resumen(wb):
    ws = wb.create_sheet("Resumen")
    _anchos(ws, [22, 12, 4, 22, 12])

    # Conteo por estado de los ítems de roadmap (Rutas + Tareas + Subtareas).
    # Se excluyen las filas de Fase, que son resúmenes estructurales.
    por_estado = {}
    por_fase = {}
    for fila in data.ROADMAP:
        if fila["Tipo"] != "Fase":
            por_estado[fila["Estado"]] = por_estado.get(fila["Estado"], 0) + 1
        else:
            por_fase[fila["Fase"]] = fila["Estado"]

    ws["A1"] = "Resumen del roadmap"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)

    _cab(ws, 3, ["Estado", "Ítems"])
    r = 4
    for estado in ["Cerrada", "En curso", "Siguiente", "Planificada", "Aplazada", "Diferida", "Backlog"]:
        n = por_estado.get(estado, 0)
        ce = ws.cell(row=r, column=1, value=estado)
        ce.border = BORDE_FINO
        ce.alignment = IZQ
        color = COLOR_ESTADO.get(estado)
        if color:
            ce.fill = PatternFill("solid", fgColor=color)
        cn = ws.cell(row=r, column=2, value=n)
        cn.border = BORDE_FINO
        cn.alignment = CENTRO
        r += 1
    total = sum(por_estado.values())
    ct = ws.cell(row=r, column=1, value="TOTAL (ítems de roadmap)")
    ct.font = Font(bold=True)
    ct.border = BORDE_FINO
    cv = ws.cell(row=r, column=2, value=total)
    cv.font = Font(bold=True)
    cv.border = BORDE_FINO
    cv.alignment = CENTRO

    # Estado por fase (cabeceras en columnas D y E).
    ws.cell(row=3, column=4).value = "Fase"
    ws.cell(row=3, column=5).value = "Estado"
    for c in (4, 5):
        cel = ws.cell(row=3, column=c)
        cel.font = FUENTE_CAB
        cel.fill = RELLENO_CAB
        cel.alignment = CENTRO
        cel.border = BORDE_FINO
    r = 4
    for fase, estado in por_fase.items():
        cf = ws.cell(row=r, column=4, value=fase)
        cf.border = BORDE_FINO
        cf.alignment = IZQ
        ce = ws.cell(row=r, column=5, value=estado)
        ce.border = BORDE_FINO
        ce.alignment = CENTRO
        color = COLOR_ESTADO.get(estado)
        if color:
            ce.fill = PatternFill("solid", fgColor=color)
        r += 1


def main():
    wb = Workbook()
    hoja_leeme(wb)
    hoja_roadmap(wb)
    hoja_gates(wb)
    hoja_perfiles(wb)
    hoja_decisiones(wb)
    hoja_resumen(wb)

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SALIDA)
    print(f"OK -> {SALIDA.relative_to(REPO_ROOT)}")
    print(f"   Filas de roadmap: {len(data.ROADMAP)} · Gates: {len(data.GATES)}")


if __name__ == "__main__":
    main()
